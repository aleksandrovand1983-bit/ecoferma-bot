import sqlite3
import io
from datetime import datetime, date, timedelta
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application, CommandHandler, CallbackQueryHandler,
    MessageHandler, filters, ContextTypes, ConversationHandler
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

import os
TOKEN = os.environ.get("BOT_TOKEN", "7819339620:AAGrMb_9p_qfZ5G1suv0Y_hcqup90BRhtbE")
DB_FILE = os.environ.get("DB_FILE", "farm.db")

# Состояния диалога
(
    MENU,
    FEEDING_COW, FEEDING_TYPE, FEEDING_VOLUME,
    MILKING_COW, MILKING_VOLUME,
    HEALTH_COW, HEALTH_NOTE,
    ADD_COW_NAME,
    DELETE_COW,
    IMPORT_COWS,
    DEATH_COW, DEATH_REASON,
    VAC_COW, VAC_NAME, VAC_NEXT,
    CALVING_COW, CALVING_COUNT, CALVING_NOTES,
    SET_MILK_PRICE,
    SET_FEED_PRICE_TYPE, SET_FEED_PRICE_VALUE,
) = range(22)


# ───────────────────────── База данных ─────────────────────────

def init_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS cows (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE NOT NULL
    )""")
    c.execute("""CREATE TABLE IF NOT EXISTS feedings (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cow_name TEXT NOT NULL,
        feed_type TEXT,
        volume REAL,
        feeding_time TEXT,
        date TEXT
    )""")
    c.execute("""CREATE TABLE IF NOT EXISTS milkings (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cow_name TEXT NOT NULL,
        volume REAL,
        milking_time TEXT,
        date TEXT
    )""")
    c.execute("""CREATE TABLE IF NOT EXISTS health_notes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cow_name TEXT NOT NULL,
        note TEXT,
        created_at TEXT
    )""")
    c.execute("""CREATE TABLE IF NOT EXISTS deaths (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cow_name TEXT NOT NULL,
        reason TEXT,
        created_at TEXT
    )""")
    c.execute("""CREATE TABLE IF NOT EXISTS vaccinations (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cow_name TEXT NOT NULL,
        vaccine_name TEXT,
        vac_date TEXT,
        next_date TEXT
    )""")
    c.execute("""CREATE TABLE IF NOT EXISTS calvings (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cow_name TEXT NOT NULL,
        calving_date TEXT,
        calves_count INTEGER,
        notes TEXT
    )""")
    c.execute("""CREATE TABLE IF NOT EXISTS settings (
        key TEXT PRIMARY KEY,
        value TEXT
    )""")
    conn.commit()
    conn.close()

def get_cows():
    conn = sqlite3.connect(DB_FILE)
    rows = conn.execute("SELECT name FROM cows ORDER BY name").fetchall()
    conn.close()
    return [r[0] for r in rows]

def save_vaccination(cow_name, vaccine_name, next_date):
    today = date.today().strftime("%Y-%m-%d")
    conn = sqlite3.connect(DB_FILE)
    conn.execute(
        "INSERT INTO vaccinations (cow_name, vaccine_name, vac_date, next_date) VALUES (?,?,?,?)",
        (cow_name, vaccine_name, today, next_date)
    )
    conn.commit()
    conn.close()

def save_calving(cow_name, calves_count, notes):
    today = date.today().strftime("%Y-%m-%d")
    conn = sqlite3.connect(DB_FILE)
    conn.execute(
        "INSERT INTO calvings (cow_name, calving_date, calves_count, notes) VALUES (?,?,?,?)",
        (cow_name, today, calves_count, notes)
    )
    conn.commit()
    conn.close()

def get_vaccinations(date_from, date_to):
    conn = sqlite3.connect(DB_FILE)
    rows = conn.execute(
        "SELECT cow_name, vaccine_name, vac_date, next_date FROM vaccinations WHERE vac_date BETWEEN ? AND ? ORDER BY vac_date",
        (date_from, date_to)
    ).fetchall()
    conn.close()
    return rows

def get_calvings(date_from, date_to):
    conn = sqlite3.connect(DB_FILE)
    rows = conn.execute(
        "SELECT cow_name, calving_date, calves_count, notes FROM calvings WHERE calving_date BETWEEN ? AND ? ORDER BY calving_date",
        (date_from, date_to)
    ).fetchall()
    conn.close()
    return rows

def get_setting(key, default="0"):
    conn = sqlite3.connect(DB_FILE)
    row = conn.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
    conn.close()
    return row[0] if row else default

def set_setting(key, value):
    conn = sqlite3.connect(DB_FILE)
    conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?,?)", (key, str(value)))
    conn.commit()
    conn.close()

def get_all_feed_prices():
    conn = sqlite3.connect(DB_FILE)
    rows = conn.execute("SELECT key, value FROM settings WHERE key LIKE 'feed_price_%'").fetchall()
    conn.close()
    return {r[0].replace("feed_price_", ""): float(r[1]) for r in rows}

def save_death(cow_name, reason):
    now = datetime.now()
    conn = sqlite3.connect(DB_FILE)
    conn.execute(
        "INSERT INTO deaths (cow_name, reason, created_at) VALUES (?,?,?)",
        (cow_name, reason, now.strftime("%Y-%m-%d %H:%M"))
    )
    conn.execute("DELETE FROM cows WHERE name=?", (cow_name,))
    conn.commit()
    conn.close()

def delete_cow(name):
    conn = sqlite3.connect(DB_FILE)
    conn.execute("DELETE FROM cows WHERE name=?", (name,))
    conn.commit()
    conn.close()

def add_cow(name):
    conn = sqlite3.connect(DB_FILE)
    try:
        conn.execute("INSERT INTO cows (name) VALUES (?)", (name,))
        conn.commit()
        return True
    except sqlite3.IntegrityError:
        return False
    finally:
        conn.close()

def save_feeding(cow_name, feed_type, volume):
    now = datetime.now()
    conn = sqlite3.connect(DB_FILE)
    conn.execute(
        "INSERT INTO feedings (cow_name, feed_type, volume, feeding_time, date) VALUES (?,?,?,?,?)",
        (cow_name, feed_type, volume, now.strftime("%H:%M"), now.strftime("%Y-%m-%d"))
    )
    conn.commit()
    conn.close()

def save_milking(cow_name, volume):
    now = datetime.now()
    conn = sqlite3.connect(DB_FILE)
    conn.execute(
        "INSERT INTO milkings (cow_name, volume, milking_time, date) VALUES (?,?,?,?)",
        (cow_name, volume, now.strftime("%H:%M"), now.strftime("%Y-%m-%d"))
    )
    conn.commit()
    conn.close()

def save_health(cow_name, note):
    now = datetime.now()
    conn = sqlite3.connect(DB_FILE)
    conn.execute(
        "INSERT INTO health_notes (cow_name, note, created_at) VALUES (?,?,?)",
        (cow_name, note, now.strftime("%Y-%m-%d %H:%M"))
    )
    conn.commit()
    conn.close()

def get_report(target_date: str):
    conn = sqlite3.connect(DB_FILE)

    feedings = conn.execute(
        "SELECT cow_name, feed_type, volume, feeding_time FROM feedings WHERE date=? ORDER BY feeding_time",
        (target_date,)
    ).fetchall()

    milkings = conn.execute(
        "SELECT cow_name, volume, milking_time FROM milkings WHERE date=? ORDER BY milking_time",
        (target_date,)
    ).fetchall()

    health = conn.execute(
        "SELECT cow_name, note, created_at FROM health_notes WHERE date(created_at)=? ORDER BY created_at",
        (target_date,)
    ).fetchall()

    conn.close()
    return feedings, milkings, health

def get_month_dates(year: int, month: int):
    import calendar
    last_day = calendar.monthrange(year, month)[1]
    date_from = f"{year}-{month:02d}-01"
    date_to = f"{year}-{month:02d}-{last_day:02d}"
    return date_from, date_to

def get_month_feed_by_cow(date_from, date_to):
    conn = sqlite3.connect(DB_FILE)
    rows = conn.execute(
        "SELECT cow_name, feed_type, SUM(volume) FROM feedings WHERE date BETWEEN ? AND ? GROUP BY cow_name, feed_type ORDER BY cow_name",
        (date_from, date_to)
    ).fetchall()
    conn.close()
    return rows

def get_month_milk_by_cow(date_from, date_to):
    conn = sqlite3.connect(DB_FILE)
    rows = conn.execute(
        "SELECT cow_name, SUM(volume) FROM milkings WHERE date BETWEEN ? AND ? GROUP BY cow_name ORDER BY cow_name",
        (date_from, date_to)
    ).fetchall()
    conn.close()
    return rows

def get_daily_totals(date_from, date_to):
    conn = sqlite3.connect(DB_FILE)
    feed = conn.execute(
        "SELECT date, SUM(volume) FROM feedings WHERE date BETWEEN ? AND ? GROUP BY date ORDER BY date",
        (date_from, date_to)
    ).fetchall()
    milk = conn.execute(
        "SELECT date, SUM(volume) FROM milkings WHERE date BETWEEN ? AND ? GROUP BY date ORDER BY date",
        (date_from, date_to)
    ).fetchall()
    conn.close()
    return feed, milk

def get_month_feedings(date_from, date_to):
    conn = sqlite3.connect(DB_FILE)
    rows = conn.execute(
        "SELECT date, cow_name, feed_type, volume, feeding_time FROM feedings WHERE date BETWEEN ? AND ? ORDER BY date, feeding_time",
        (date_from, date_to)
    ).fetchall()
    conn.close()
    return rows

def get_month_milkings(date_from, date_to):
    conn = sqlite3.connect(DB_FILE)
    rows = conn.execute(
        "SELECT date, cow_name, volume, milking_time FROM milkings WHERE date BETWEEN ? AND ? ORDER BY date, milking_time",
        (date_from, date_to)
    ).fetchall()
    conn.close()
    return rows

def get_month_health(date_from, date_to):
    conn = sqlite3.connect(DB_FILE)
    rows = conn.execute(
        "SELECT date(created_at), cow_name, note, created_at FROM health_notes WHERE date(created_at) BETWEEN ? AND ? ORDER BY created_at",
        (date_from, date_to)
    ).fetchall()
    conn.close()
    return rows

def get_deaths(date_from: str, date_to: str):
    conn = sqlite3.connect(DB_FILE)
    rows = conn.execute(
        "SELECT cow_name, reason, created_at FROM deaths WHERE date(created_at) BETWEEN ? AND ? ORDER BY created_at",
        (date_from, date_to)
    ).fetchall()
    conn.close()
    return rows


# ───────────────────────── Клавиатуры ─────────────────────────

def main_menu():
    kb = [
        [InlineKeyboardButton("🥗 Кормление", callback_data="feeding"),
         InlineKeyboardButton("🥛 Удой", callback_data="milking")],
        [InlineKeyboardButton("🏥 Здоровье/Заметка", callback_data="health"),
         InlineKeyboardButton("📊 Отчёт сегодня", callback_data="report_today")],
        [InlineKeyboardButton("📅 Отчёт за 7 дней", callback_data="report_week"),
         InlineKeyboardButton("🐄 Добавить корову", callback_data="add_cow")],
        [InlineKeyboardButton("📥 Excel-отчёт за сегодня", callback_data="excel_today"),
         InlineKeyboardButton("📥 Excel за 7 дней", callback_data="excel_week")],
        [InlineKeyboardButton("📊 Аналитика за месяц (Excel)", callback_data="excel_month")],
        [InlineKeyboardButton("🗑 Удалить корову", callback_data="delete_cow")],
        [InlineKeyboardButton("📋 Загрузить список коров", callback_data="import_cows")],
        [InlineKeyboardButton("💀 Падёж коровы", callback_data="death_cow")],
        [InlineKeyboardButton("🐾 Ветучёт", callback_data="vet_menu"),
         InlineKeyboardButton("💰 Финансы", callback_data="finance_menu")],
    ]
    return InlineKeyboardMarkup(kb)

def cows_keyboard(action_prefix):
    cows = get_cows()
    if not cows:
        return None
    kb = [[InlineKeyboardButton(c, callback_data=f"{action_prefix}:{c}")] for c in cows]
    kb.append([InlineKeyboardButton("◀️ Отмена", callback_data="cancel")])
    return InlineKeyboardMarkup(kb)

def feed_types_keyboard():
    types = ["Сено", "Силос", "Зерно", "Комбикорм", "Трава", "Другое"]
    kb = [[InlineKeyboardButton(t, callback_data=f"feedtype:{t}")] for t in types]
    kb.append([InlineKeyboardButton("◀️ Отмена", callback_data="cancel")])
    return InlineKeyboardMarkup(kb)

def death_reasons_keyboard():
    reasons = ["Болезнь", "Травма", "Осложнения при родах", "Отравление", "Старость", "Другое"]
    kb = [[InlineKeyboardButton(r, callback_data=f"dreason:{r}")] for r in reasons]
    kb.append([InlineKeyboardButton("◀️ Отмена", callback_data="cancel")])
    return InlineKeyboardMarkup(kb)

def vet_menu_keyboard():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("💉 Вакцинация", callback_data="vac_start"),
         InlineKeyboardButton("🐮 Отёл", callback_data="calving_start")],
        [InlineKeyboardButton("◀️ Главное меню", callback_data="cancel")],
    ])

def finance_menu_keyboard():
    milk_price = get_setting("milk_price", "0")
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(f"🥛 Цена молока: {milk_price} руб/л", callback_data="set_milk_price")],
        [InlineKeyboardButton("🥗 Цены на корма", callback_data="set_feed_price")],
        [InlineKeyboardButton("📊 Финотчёт за сегодня", callback_data="finance_today"),
         InlineKeyboardButton("📊 Финотчёт за месяц", callback_data="finance_month")],
        [InlineKeyboardButton("◀️ Главное меню", callback_data="cancel")],
    ])

def feed_types_price_keyboard():
    types = ["Сено", "Силос", "Зерно", "Комбикорм", "Трава", "Другое"]
    prices = get_all_feed_prices()
    kb = []
    for t in types:
        price = prices.get(t, 0)
        kb.append([InlineKeyboardButton(f"{t}: {price} руб/кг", callback_data=f"fptype:{t}")])
    kb.append([InlineKeyboardButton("◀️ Отмена", callback_data="cancel")])
    return InlineKeyboardMarkup(kb)

def cancel_keyboard():
    return InlineKeyboardMarkup([[InlineKeyboardButton("◀️ Отмена", callback_data="cancel")]])


# ───────────────────────── Хендлеры ─────────────────────────

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg = update.message or update.callback_query.message
    await msg.reply_text(
        "🐄 *Фермерский журнал*\n\nВыберите действие:",
        reply_markup=main_menu(),
        parse_mode="Markdown"
    )
    return MENU

async def menu_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data

    if data == "cancel":
        await query.edit_message_text("Главное меню:", reply_markup=main_menu())
        return MENU

    # ── Кормление ──
    if data == "feeding":
        kb = cows_keyboard("feed_cow")
        if not kb:
            await query.edit_message_text(
                "Сначала добавьте корову через кнопку 🐄 Добавить корову",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️ Назад", callback_data="cancel")]])
            )
            return MENU
        await query.edit_message_text("Выберите корову:", reply_markup=kb)
        return FEEDING_COW

    # ── Удой ──
    if data == "milking":
        kb = cows_keyboard("milk_cow")
        if not kb:
            await query.edit_message_text(
                "Сначала добавьте корову через кнопку 🐄 Добавить корову",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️ Назад", callback_data="cancel")]])
            )
            return MENU
        await query.edit_message_text("Выберите корову:", reply_markup=kb)
        return MILKING_COW

    # ── Здоровье ──
    if data == "health":
        kb = cows_keyboard("health_cow")
        if not kb:
            await query.edit_message_text(
                "Сначала добавьте корову через кнопку 🐄 Добавить корову",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️ Назад", callback_data="cancel")]])
            )
            return MENU
        await query.edit_message_text("Выберите корову для заметки:", reply_markup=kb)
        return HEALTH_COW

    # ── Ветучёт ──
    if data == "vet_menu":
        return await vet_menu(update, context)
    if data == "vac_start":
        return await vac_start(update, context)
    if data == "calving_start":
        return await calving_start(update, context)

    # ── Финансы ──
    if data == "finance_menu":
        return await finance_menu(update, context)
    if data == "set_milk_price":
        return await set_milk_price_start(update, context)
    if data == "set_feed_price":
        return await set_feed_price_start(update, context)
    if data == "finance_today":
        return await finance_today(update, context)
    if data == "finance_month":
        return await finance_month(update, context)

    # ── Падёж коровы ──
    if data == "death_cow":
        return await death_cow_start(update, context)

    # ── Загрузить список коров ──
    if data == "import_cows":
        return await import_cows_start(update, context)

    # ── Удалить корову ──
    if data == "delete_cow":
        return await delete_cow_start(update, context)

    # ── Добавить корову ──
    if data == "add_cow":
        await query.edit_message_text("Введите имя коровы:", reply_markup=cancel_keyboard())
        return ADD_COW_NAME

    # ── Отчёт сегодня ──
    if data == "report_today":
        today = date.today().strftime("%Y-%m-%d")
        text = build_report(today, "Сегодня")
        await query.edit_message_text(text, reply_markup=main_menu(), parse_mode="Markdown")
        return MENU

    # ── Отчёт 7 дней ──
    if data == "report_week":
        lines = []
        for i in range(6, -1, -1):
            d = (date.today() - timedelta(days=i)).strftime("%Y-%m-%d")
            label = d[8:10] + "." + d[5:7]
            lines.append(build_report(d, label))
        text = "\n\n".join(lines)
        await query.edit_message_text(text[:4000], reply_markup=main_menu(), parse_mode="Markdown")
        return MENU

    return MENU


# ── Кормление: выбор коровы ──
async def feeding_cow(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "cancel":
        await query.edit_message_text("Главное меню:", reply_markup=main_menu())
        return MENU
    cow = query.data.split(":", 1)[1]
    context.user_data["feeding_cow"] = cow
    await query.edit_message_text(f"Корова: *{cow}*\nВыберите тип корма:", reply_markup=feed_types_keyboard(), parse_mode="Markdown")
    return FEEDING_TYPE

# ── Кормление: выбор типа корма ──
async def feeding_type(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "cancel":
        await query.edit_message_text("Главное меню:", reply_markup=main_menu())
        return MENU
    feed_type = query.data.split(":", 1)[1]
    context.user_data["feeding_type"] = feed_type
    cow = context.user_data["feeding_cow"]
    await query.edit_message_text(
        f"Корова: *{cow}*\nКорм: *{feed_type}*\n\nВведите объём (кг):",
        reply_markup=cancel_keyboard(),
        parse_mode="Markdown"
    )
    return FEEDING_VOLUME

# ── Кормление: ввод объёма ──
async def feeding_volume(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip().replace(",", ".")
    try:
        volume = float(text)
    except ValueError:
        await update.message.reply_text("Введите число, например: 5 или 3.5", reply_markup=cancel_keyboard())
        return FEEDING_VOLUME

    cow = context.user_data["feeding_cow"]
    feed_type = context.user_data["feeding_type"]
    save_feeding(cow, feed_type, volume)
    now = datetime.now().strftime("%H:%M")
    await update.message.reply_text(
        f"✅ Записано!\n🐄 *{cow}* — {feed_type}, {volume} кг в {now}",
        reply_markup=main_menu(),
        parse_mode="Markdown"
    )
    return MENU


# ── Удой: выбор коровы ──
async def milking_cow(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "cancel":
        await query.edit_message_text("Главное меню:", reply_markup=main_menu())
        return MENU
    cow = query.data.split(":", 1)[1]
    context.user_data["milking_cow"] = cow
    await query.edit_message_text(
        f"Корова: *{cow}*\n\nВведите надой (литры):",
        reply_markup=cancel_keyboard(),
        parse_mode="Markdown"
    )
    return MILKING_VOLUME

# ── Удой: ввод объёма ──
async def milking_volume(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip().replace(",", ".")
    try:
        volume = float(text)
    except ValueError:
        await update.message.reply_text("Введите число, например: 12 или 8.5", reply_markup=cancel_keyboard())
        return MILKING_VOLUME

    cow = context.user_data["milking_cow"]
    save_milking(cow, volume)
    now = datetime.now().strftime("%H:%M")
    await update.message.reply_text(
        f"✅ Записано!\n🥛 *{cow}* — надой {volume} л в {now}",
        reply_markup=main_menu(),
        parse_mode="Markdown"
    )
    return MENU


# ── Здоровье: выбор коровы ──
async def health_cow(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "cancel":
        await query.edit_message_text("Главное меню:", reply_markup=main_menu())
        return MENU
    cow = query.data.split(":", 1)[1]
    context.user_data["health_cow"] = cow
    await query.edit_message_text(
        f"Корова: *{cow}*\n\nНапишите заметку о здоровье или наблюдение:",
        reply_markup=cancel_keyboard(),
        parse_mode="Markdown"
    )
    return HEALTH_NOTE

# ── Здоровье: ввод заметки ──
async def health_note(update: Update, context: ContextTypes.DEFAULT_TYPE):
    note = update.message.text.strip()
    cow = context.user_data["health_cow"]
    save_health(cow, note)
    await update.message.reply_text(
        f"✅ Заметка сохранена!\n🏥 *{cow}*: {note}",
        reply_markup=main_menu(),
        parse_mode="Markdown"
    )
    return MENU


# ── Добавить корову ──
async def add_cow_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    name = update.message.text.strip()
    if not name:
        await update.message.reply_text("Имя не может быть пустым. Попробуйте ещё раз:")
        return ADD_COW_NAME
    if add_cow(name):
        await update.message.reply_text(
            f"✅ Корова *{name}* добавлена!",
            reply_markup=main_menu(),
            parse_mode="Markdown"
        )
    else:
        await update.message.reply_text(
            f"Корова с именем *{name}* уже есть.",
            reply_markup=main_menu(),
            parse_mode="Markdown"
        )
    return MENU


# ───────────────────────── Отчёт ─────────────────────────

def build_report(target_date: str, label: str) -> str:
    feedings, milkings, health = get_report(target_date)
    lines = [f"📋 *{label}* ({target_date})"]

    if feedings:
        lines.append("\n🥗 *Кормления:*")
        for cow, ftype, vol, ftime in feedings:
            lines.append(f"  • {ftime} — {cow}: {ftype}, {vol} кг")
        total_feed = sum(r[2] for r in feedings)
        lines.append(f"  Итого: {total_feed} кг")
    else:
        lines.append("\n🥗 Кормлений нет")

    if milkings:
        lines.append("\n🥛 *Удои:*")
        for cow, vol, mtime in milkings:
            lines.append(f"  • {mtime} — {cow}: {vol} л")
        total_milk = sum(r[1] for r in milkings)
        lines.append(f"  Итого: {total_milk} л")
    else:
        lines.append("\n🥛 Удоев нет")

    if health:
        lines.append("\n🏥 *Заметки:*")
        for cow, note, created in health:
            lines.append(f"  • {cow}: {note}")

    return "\n".join(lines)


# ───────────────────────── Падёж коровы ─────────────────────────

async def death_cow_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    kb = cows_keyboard("death_cow_sel")
    if not kb:
        await query.edit_message_text(
            "Список коров пуст.",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️ Назад", callback_data="cancel")]])
        )
        return MENU
    await query.edit_message_text("Выберите корову:", reply_markup=kb)
    return DEATH_COW

async def death_cow_select(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "cancel":
        await query.edit_message_text("Главное меню:", reply_markup=main_menu())
        return MENU
    cow = query.data.split(":", 1)[1]
    context.user_data["death_cow"] = cow
    await query.edit_message_text(
        f"Корова: *{cow}*\n\nУкажите причину падежа:",
        reply_markup=death_reasons_keyboard(),
        parse_mode="Markdown"
    )
    return DEATH_REASON

async def death_reason_select(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "cancel":
        await query.edit_message_text("Главное меню:", reply_markup=main_menu())
        return MENU
    reason = query.data.split(":", 1)[1]
    cow = context.user_data["death_cow"]
    save_death(cow, reason)
    now = datetime.now().strftime("%d.%m.%Y %H:%M")
    await query.edit_message_text(
        f"📋 Падёж зафиксирован\n\n🐄 Корова: *{cow}*\nПричина: *{reason}*\nДата: {now}\n\nКорова удалена из активного списка.",
        reply_markup=main_menu(),
        parse_mode="Markdown"
    )
    return MENU


# ───────────────────────── Импорт коров из файла ─────────────────────────

async def import_cows_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.edit_message_text(
        "Отправьте файл со списком коров:\n\n"
        "• *.txt* — одно имя на строку\n"
        "• *.xlsx* — имена в первом столбце\n\n"
        "Дубликаты будут пропущены автоматически.",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️ Отмена", callback_data="cancel")]])
    )
    return IMPORT_COWS

async def import_cows_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    doc = update.message.document
    if not doc:
        await update.message.reply_text("Пожалуйста, отправьте файл .txt или .xlsx.")
        return IMPORT_COWS

    fname = doc.file_name.lower()
    if not (fname.endswith(".txt") or fname.endswith(".xlsx")):
        await update.message.reply_text("Поддерживаются только .txt и .xlsx файлы.")
        return IMPORT_COWS

    file = await doc.get_file()
    data = await file.download_as_bytearray()

    names = []
    if fname.endswith(".txt"):
        text = data.decode("utf-8", errors="ignore")
        names = [line.strip() for line in text.splitlines() if line.strip()]
    else:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
            val = row[0]
            if val and str(val).strip():
                names.append(str(val).strip())

    added, skipped = 0, 0
    for name in names:
        if add_cow(name):
            added += 1
        else:
            skipped += 1

    await update.message.reply_text(
        f"✅ Загрузка завершена!\n\n"
        f"Добавлено коров: *{added}*\n"
        f"Пропущено (уже есть): *{skipped}*",
        reply_markup=main_menu(),
        parse_mode="Markdown"
    )
    return MENU


# ───────────────────────── Удаление коровы ─────────────────────────

async def delete_cow_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    kb = cows_keyboard("del_cow")
    if not kb:
        await query.edit_message_text(
            "Список коров пуст.",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️ Назад", callback_data="cancel")]])
        )
        return MENU
    await query.edit_message_text("Выберите корову для удаления:", reply_markup=kb)
    return DELETE_COW

async def delete_cow_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "cancel":
        await query.edit_message_text("Главное меню:", reply_markup=main_menu())
        return MENU
    cow = query.data.split(":", 1)[1]
    delete_cow(cow)
    await query.edit_message_text(
        f"✅ Корова *{cow}* удалена.",
        reply_markup=main_menu(),
        parse_mode="Markdown"
    )
    return MENU


# ───────────────────────── Ветучёт ─────────────────────────

async def vet_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.edit_message_text("🐾 *Ветеринарный учёт*\n\nВыберите действие:", reply_markup=vet_menu_keyboard(), parse_mode="Markdown")
    return MENU

async def vac_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    kb = cows_keyboard("vac_cow")
    if not kb:
        await query.edit_message_text("Список коров пуст.", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️ Назад", callback_data="cancel")]]))
        return MENU
    await query.edit_message_text("💉 Выберите корову для вакцинации:", reply_markup=kb)
    return VAC_COW

async def vac_cow_select(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "cancel":
        await query.edit_message_text("Главное меню:", reply_markup=main_menu())
        return MENU
    cow = query.data.split(":", 1)[1]
    context.user_data["vac_cow"] = cow
    await query.edit_message_text(f"Корова: *{cow}*\n\nВведите название вакцины:", reply_markup=cancel_keyboard(), parse_mode="Markdown")
    return VAC_NAME

async def vac_name_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["vac_name"] = update.message.text.strip()
    cow = context.user_data["vac_cow"]
    await update.message.reply_text(
        f"Корова: *{cow}*\nВакцина: *{context.user_data['vac_name']}*\n\nВведите дату следующей вакцинации (например: 2025-12-01 или «через 6 мес»):",
        reply_markup=cancel_keyboard(), parse_mode="Markdown"
    )
    return VAC_NEXT

async def vac_next_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    next_date = update.message.text.strip()
    cow = context.user_data["vac_cow"]
    vaccine = context.user_data["vac_name"]
    save_vaccination(cow, vaccine, next_date)
    today = date.today().strftime("%d.%m.%Y")
    await update.message.reply_text(
        f"✅ Вакцинация записана!\n\n💉 *{cow}* — {vaccine}\nДата: {today}\nСледующая: {next_date}",
        reply_markup=main_menu(), parse_mode="Markdown"
    )
    return MENU

async def calving_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    kb = cows_keyboard("calving_cow")
    if not kb:
        await query.edit_message_text("Список коров пуст.", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("◀️ Назад", callback_data="cancel")]]))
        return MENU
    await query.edit_message_text("🐮 Выберите корову:", reply_markup=kb)
    return CALVING_COW

async def calving_cow_select(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "cancel":
        await query.edit_message_text("Главное меню:", reply_markup=main_menu())
        return MENU
    cow = query.data.split(":", 1)[1]
    context.user_data["calving_cow"] = cow
    await query.edit_message_text(f"Корова: *{cow}*\n\nСколько телят?", reply_markup=cancel_keyboard(), parse_mode="Markdown")
    return CALVING_COUNT

async def calving_count_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if not text.isdigit():
        await update.message.reply_text("Введите число, например: 1", reply_markup=cancel_keyboard())
        return CALVING_COUNT
    context.user_data["calving_count"] = int(text)
    cow = context.user_data["calving_cow"]
    await update.message.reply_text(
        f"Корова: *{cow}*, телят: *{text}*\n\nДобавьте заметку (или напишите «-» если нет):",
        reply_markup=cancel_keyboard(), parse_mode="Markdown"
    )
    return CALVING_NOTES

async def calving_notes_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    notes = update.message.text.strip()
    if notes == "-":
        notes = ""
    cow = context.user_data["calving_cow"]
    count = context.user_data["calving_count"]
    save_calving(cow, count, notes)
    today = date.today().strftime("%d.%m.%Y")
    await update.message.reply_text(
        f"✅ Отёл записан!\n\n🐮 *{cow}* — {count} телёнок(телят)\nДата: {today}" + (f"\nЗаметка: {notes}" if notes else ""),
        reply_markup=main_menu(), parse_mode="Markdown"
    )
    return MENU


# ───────────────────────── Финансы ─────────────────────────

async def finance_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.edit_message_text("💰 *Финансовый учёт*\n\nВыберите действие:", reply_markup=finance_menu_keyboard(), parse_mode="Markdown")
    return MENU

async def set_milk_price_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    current = get_setting("milk_price", "0")
    await query.edit_message_text(f"Текущая цена молока: *{current} руб/л*\n\nВведите новую цену (руб за литр):", reply_markup=cancel_keyboard(), parse_mode="Markdown")
    return SET_MILK_PRICE

async def set_milk_price_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip().replace(",", ".")
    try:
        price = float(text)
    except ValueError:
        await update.message.reply_text("Введите число, например: 45 или 45.50", reply_markup=cancel_keyboard())
        return SET_MILK_PRICE
    set_setting("milk_price", price)
    await update.message.reply_text(f"✅ Цена молока установлена: *{price} руб/л*", reply_markup=main_menu(), parse_mode="Markdown")
    return MENU

async def set_feed_price_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.edit_message_text("Выберите тип корма для установки цены:", reply_markup=feed_types_price_keyboard())
    return SET_FEED_PRICE_TYPE

async def set_feed_price_type(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "cancel":
        await query.edit_message_text("Главное меню:", reply_markup=main_menu())
        return MENU
    ftype = query.data.split(":", 1)[1]
    context.user_data["feed_price_type"] = ftype
    current = get_setting(f"feed_price_{ftype}", "0")
    await query.edit_message_text(
        f"Корм: *{ftype}*\nТекущая цена: *{current} руб/кг*\n\nВведите новую цену (руб за кг):",
        reply_markup=cancel_keyboard(), parse_mode="Markdown"
    )
    return SET_FEED_PRICE_VALUE

async def set_feed_price_value(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip().replace(",", ".")
    try:
        price = float(text)
    except ValueError:
        await update.message.reply_text("Введите число, например: 8 или 12.50", reply_markup=cancel_keyboard())
        return SET_FEED_PRICE_VALUE
    ftype = context.user_data["feed_price_type"]
    set_setting(f"feed_price_{ftype}", price)
    await update.message.reply_text(f"✅ Цена *{ftype}*: *{price} руб/кг*", reply_markup=main_menu(), parse_mode="Markdown")
    return MENU

def build_finance_report(date_from: str, date_to: str) -> str:
    milk_price = float(get_setting("milk_price", "0"))
    feed_prices = get_all_feed_prices()

    conn = sqlite3.connect(DB_FILE)
    total_milk = conn.execute(
        "SELECT COALESCE(SUM(volume),0) FROM milkings WHERE date BETWEEN ? AND ?", (date_from, date_to)
    ).fetchone()[0]
    feed_rows = conn.execute(
        "SELECT feed_type, COALESCE(SUM(volume),0) FROM feedings WHERE date BETWEEN ? AND ? GROUP BY feed_type", (date_from, date_to)
    ).fetchall()
    conn.close()

    milk_income = round(total_milk * milk_price, 2)
    feed_cost = 0.0
    feed_lines = []
    for ftype, vol in feed_rows:
        price = feed_prices.get(ftype, 0)
        cost = round(vol * price, 2)
        feed_cost += cost
        feed_lines.append(f"  • {ftype}: {vol} кг × {price} руб = {cost} руб")

    feed_cost = round(feed_cost, 2)
    profit = round(milk_income - feed_cost, 2)

    lines = [
        f"💰 *Финансовый отчёт* ({date_from} — {date_to})\n",
        f"🥛 Молоко: {total_milk} л × {milk_price} руб = *{milk_income} руб*\n",
        "🥗 *Расход на корма:*",
    ] + (feed_lines if feed_lines else ["  Нет данных"]) + [
        f"\n  Итого кормов: *{feed_cost} руб*",
        f"\n📈 Прибыль: *{profit} руб*",
    ]
    return "\n".join(lines)

async def finance_today(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    today = date.today().strftime("%Y-%m-%d")
    text = build_finance_report(today, today)
    await query.edit_message_text(text, reply_markup=finance_menu_keyboard(), parse_mode="Markdown")
    return MENU

async def finance_month(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    today = date.today()
    date_from, date_to = get_month_dates(today.year, today.month)
    text = build_finance_report(date_from, date_to)
    await query.edit_message_text(text, reply_markup=finance_menu_keyboard(), parse_mode="Markdown")
    return MENU


# ───────────────────────── Excel ─────────────────────────

def _style_header(ws, headers, fill_color):
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(bold=True, color="FFFFFF")
    for col, title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

def build_excel(dates: list) -> io.BytesIO:
    wb = Workbook()

    # Лист Кормления
    ws_f = wb.active
    ws_f.title = "Кормления"
    _style_header(ws_f, ["Дата", "Корова", "Корм", "Объём (кг)", "Время"], "4CAF50")
    for d in dates:
        feedings, _, _ = get_report(d)
        for cow, ftype, vol, ftime in feedings:
            ws_f.append([d, cow, ftype, vol, ftime])

    # Лист Удои
    ws_m = wb.create_sheet("Удои")
    _style_header(ws_m, ["Дата", "Корова", "Надой (л)", "Время"], "2196F3")
    for d in dates:
        _, milkings, _ = get_report(d)
        for cow, vol, mtime in milkings:
            ws_m.append([d, cow, vol, mtime])

    # Лист Здоровье
    ws_h = wb.create_sheet("Здоровье")
    _style_header(ws_h, ["Дата", "Корова", "Заметка", "Время"], "F44336")
    for d in dates:
        _, _, health = get_report(d)
        for cow, note, created in health:
            ws_h.append([d, cow, note, created])

    # Лист Падёж
    ws_d = wb.create_sheet("Падёж")
    _style_header(ws_d, ["Дата", "Корова", "Причина", "Время"], "7B1FA2")
    deaths = get_deaths(dates[0], dates[-1])
    for cow, reason, created in deaths:
        ws_d.append([created[:10], cow, reason, created[11:]])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

async def excel_today(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    today = date.today().strftime("%Y-%m-%d")
    buf = build_excel([today])
    filename = f"ecoferma_{today}.xlsx"
    await query.message.reply_document(
        document=buf,
        filename=filename,
        caption=f"📥 Отчёт за {today}"
    )
    return MENU

async def excel_week(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    dates = [(date.today() - timedelta(days=i)).strftime("%Y-%m-%d") for i in range(6, -1, -1)]
    buf = build_excel(dates)
    today = date.today().strftime("%Y-%m-%d")
    filename = f"ecoferma_7days_{today}.xlsx"
    await query.message.reply_document(
        document=buf,
        filename=filename,
        caption="📥 Отчёт за последние 7 дней"
    )
    return MENU

def build_excel_month(year: int, month: int) -> io.BytesIO:
    import calendar
    date_from, date_to = get_month_dates(year, month)
    month_label = f"{year}-{month:02d}"
    wb = Workbook()

    # ── Лист 1: Сводка по коровам ──
    ws_sum = wb.active
    ws_sum.title = "Сводка"
    _style_header(ws_sum, ["Корова", "Корм всего (кг)", "Надой всего (л)", "Заметок здоровья", "Падёж"], "37474F")

    feed_by_cow = {}
    for cow, ftype, vol in get_month_feed_by_cow(date_from, date_to):
        feed_by_cow[cow] = feed_by_cow.get(cow, 0) + (vol or 0)

    milk_by_cow = {cow: vol for cow, vol in get_month_milk_by_cow(date_from, date_to)}

    health_rows = get_month_health(date_from, date_to)
    health_count = {}
    for _, cow, _, _ in health_rows:
        health_count[cow] = health_count.get(cow, 0) + 1

    death_rows = get_deaths(date_from, date_to)
    dead_cows = {cow for cow, _, _ in death_rows}

    all_cows = sorted(set(list(feed_by_cow) + list(milk_by_cow) + list(health_count) + list(dead_cows)))
    for cow in all_cows:
        ws_sum.append([
            cow,
            round(feed_by_cow.get(cow, 0), 2),
            round(milk_by_cow.get(cow, 0), 2),
            health_count.get(cow, 0),
            "Да" if cow in dead_cows else "Нет"
        ])

    # Итоговая строка
    last = ws_sum.max_row + 1
    ws_sum.cell(last, 1, "ИТОГО").font = Font(bold=True)
    ws_sum.cell(last, 2, round(sum(feed_by_cow.values()), 2)).font = Font(bold=True)
    ws_sum.cell(last, 3, round(sum(milk_by_cow.values()), 2)).font = Font(bold=True)
    ws_sum.cell(last, 4, sum(health_count.values())).font = Font(bold=True)
    ws_sum.cell(last, 5, len(dead_cows)).font = Font(bold=True)

    # ── Лист 2: По дням ──
    ws_day = wb.create_sheet("По дням")
    _style_header(ws_day, ["Дата", "Корм (кг)", "Надой (л)"], "00796B")

    feed_daily, milk_daily = get_daily_totals(date_from, date_to)
    feed_dict = dict(feed_daily)
    milk_dict = dict(milk_daily)
    all_dates = sorted(set(list(feed_dict) + list(milk_dict)))
    for d in all_dates:
        ws_day.append([d, round(feed_dict.get(d, 0), 2), round(milk_dict.get(d, 0), 2)])

    # Диаграмма надоев по дням
    if len(all_dates) > 0:
        chart = BarChart()
        chart.title = "Надой по дням (л)"
        chart.style = 10
        chart.y_axis.title = "Литры"
        chart.x_axis.title = "Дата"
        chart.width = 20
        chart.height = 12
        data_ref = Reference(ws_day, min_col=3, min_row=1, max_row=ws_day.max_row)
        cats = Reference(ws_day, min_col=1, min_row=2, max_row=ws_day.max_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        ws_day.add_chart(chart, f"E2")

    # ── Лист 3: Кормления ──
    ws_f = wb.create_sheet("Кормления")
    _style_header(ws_f, ["Дата", "Корова", "Корм", "Объём (кг)", "Время"], "4CAF50")
    for d, cow, ftype, vol, ftime in get_month_feedings(date_from, date_to):
        ws_f.append([d, cow, ftype, vol, ftime])

    # ── Лист 4: Удои ──
    ws_m = wb.create_sheet("Удои")
    _style_header(ws_m, ["Дата", "Корова", "Надой (л)", "Время"], "2196F3")
    for d, cow, vol, mtime in get_month_milkings(date_from, date_to):
        ws_m.append([d, cow, vol, mtime])

    # ── Лист 5: Здоровье ──
    ws_h = wb.create_sheet("Здоровье")
    _style_header(ws_h, ["Дата", "Корова", "Заметка", "Время"], "F44336")
    for d, cow, note, created in health_rows:
        ws_h.append([d, cow, note, created])

    # ── Лист 6: Падёж ──
    ws_d = wb.create_sheet("Падёж")
    _style_header(ws_d, ["Дата", "Корова", "Причина", "Время"], "7B1FA2")
    for cow, reason, created in death_rows:
        ws_d.append([created[:10], cow, reason, created[11:]])

    # ── Лист 7: Вакцинации ──
    ws_v = wb.create_sheet("Вакцинации")
    _style_header(ws_v, ["Дата вакцинации", "Корова", "Вакцина", "Следующая дата"], "0277BD")
    for cow, vaccine, vac_date, next_date in get_vaccinations(date_from, date_to):
        ws_v.append([vac_date, cow, vaccine, next_date])

    # ── Лист 8: Отёлы ──
    ws_c = wb.create_sheet("Отёлы")
    _style_header(ws_c, ["Дата", "Корова", "Телят", "Заметка"], "558B2F")
    for cow, calving_date, count, notes in get_calvings(date_from, date_to):
        ws_c.append([calving_date, cow, count, notes or ""])

    # ── Лист 9: Финансы ──
    ws_fin = wb.create_sheet("Финансы")
    _style_header(ws_fin, ["Показатель", "Значение"], "F57F17")
    milk_price = float(get_setting("milk_price", "0"))
    feed_prices = get_all_feed_prices()
    conn = sqlite3.connect(DB_FILE)
    total_milk = conn.execute(
        "SELECT COALESCE(SUM(volume),0) FROM milkings WHERE date BETWEEN ? AND ?", (date_from, date_to)
    ).fetchone()[0]
    feed_rows_fin = conn.execute(
        "SELECT feed_type, COALESCE(SUM(volume),0) FROM feedings WHERE date BETWEEN ? AND ? GROUP BY feed_type", (date_from, date_to)
    ).fetchall()
    conn.close()
    milk_income = round(total_milk * milk_price, 2)
    feed_cost = 0.0
    ws_fin.append(["Цена молока (руб/л)", milk_price])
    ws_fin.append(["Молоко продано (л)", round(total_milk, 2)])
    ws_fin.append(["Выручка от молока (руб)", milk_income])
    ws_fin.append(["", ""])
    for ftype, vol in feed_rows_fin:
        price = feed_prices.get(ftype, 0)
        cost = round(vol * price, 2)
        feed_cost += cost
        ws_fin.append([f"Корм {ftype} ({vol} кг × {price} руб/кг)", cost])
    feed_cost = round(feed_cost, 2)
    ws_fin.append(["", ""])
    ws_fin.append(["Итого расход на корма (руб)", feed_cost])
    profit_row = ws_fin.max_row + 1
    ws_fin.cell(profit_row, 1, "Прибыль (руб)").font = Font(bold=True)
    ws_fin.cell(profit_row, 2, round(milk_income - feed_cost, 2)).font = Font(bold=True)
    ws_fin.column_dimensions["A"].width = 35
    ws_fin.column_dimensions["B"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

async def excel_month(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer("Формирую отчёт...")
    today = date.today()
    buf = build_excel_month(today.year, today.month)
    filename = f"ecoferma_{today.year}_{today.month:02d}.xlsx"
    import calendar
    month_name = ["январь","февраль","март","апрель","май","июнь",
                  "июль","август","сентябрь","октябрь","ноябрь","декабрь"][today.month - 1]
    await query.message.reply_document(
        document=buf,
        filename=filename,
        caption=f"📊 Аналитика за {month_name} {today.year}"
    )
    return MENU


# ───────────────────────── Запуск ─────────────────────────

def main():
    init_db()

    app = Application.builder().token(TOKEN).build()

    conv = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            MENU: [
                CallbackQueryHandler(excel_today, pattern="^excel_today$"),
                CallbackQueryHandler(excel_week, pattern="^excel_week$"),
                CallbackQueryHandler(excel_month, pattern="^excel_month$"),
                CallbackQueryHandler(delete_cow_start, pattern="^delete_cow$"),
                CallbackQueryHandler(menu_handler),
            ],
            DELETE_COW: [CallbackQueryHandler(delete_cow_confirm)],
            IMPORT_COWS: [
                MessageHandler(filters.Document.ALL, import_cows_file),
                CallbackQueryHandler(menu_handler),
            ],
            DEATH_COW: [CallbackQueryHandler(death_cow_select)],
            DEATH_REASON: [CallbackQueryHandler(death_reason_select)],
            VAC_COW: [CallbackQueryHandler(vac_cow_select)],
            VAC_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, vac_name_input),
                       CallbackQueryHandler(menu_handler)],
            VAC_NEXT: [MessageHandler(filters.TEXT & ~filters.COMMAND, vac_next_input),
                       CallbackQueryHandler(menu_handler)],
            CALVING_COW: [CallbackQueryHandler(calving_cow_select)],
            CALVING_COUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, calving_count_input),
                            CallbackQueryHandler(menu_handler)],
            CALVING_NOTES: [MessageHandler(filters.TEXT & ~filters.COMMAND, calving_notes_input),
                            CallbackQueryHandler(menu_handler)],
            SET_MILK_PRICE: [MessageHandler(filters.TEXT & ~filters.COMMAND, set_milk_price_input),
                             CallbackQueryHandler(menu_handler)],
            SET_FEED_PRICE_TYPE: [CallbackQueryHandler(set_feed_price_type)],
            SET_FEED_PRICE_VALUE: [MessageHandler(filters.TEXT & ~filters.COMMAND, set_feed_price_value),
                                   CallbackQueryHandler(menu_handler)],
            FEEDING_COW: [CallbackQueryHandler(feeding_cow)],
            FEEDING_TYPE: [CallbackQueryHandler(feeding_type)],
            FEEDING_VOLUME: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, feeding_volume),
                CallbackQueryHandler(menu_handler),
            ],
            MILKING_COW: [CallbackQueryHandler(milking_cow)],
            MILKING_VOLUME: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, milking_volume),
                CallbackQueryHandler(menu_handler),
            ],
            HEALTH_COW: [CallbackQueryHandler(health_cow)],
            HEALTH_NOTE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, health_note),
                CallbackQueryHandler(menu_handler),
            ],
            ADD_COW_NAME: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, add_cow_name),
                CallbackQueryHandler(menu_handler),
            ],
        },
        fallbacks=[CommandHandler("start", start)],
    )

    app.add_handler(conv)
    print("Бот запущен...")
    app.run_polling()

if __name__ == "__main__":
    main()
